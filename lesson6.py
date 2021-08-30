# Author : 刘欢
# Time : 2021/8/30 0:18
# E-mail : 962171756@qq.com


# 1.分析需求
# 2.编写自动化测试用例             read_data()
# 3.发送请求，得到响应结果         func()
# 4.执行结果(响应结果) VS 预期结果
# 5.写入最终的真实结果到测试用例   write_data()
import requests
import openpyxl

def func(url,body,header):
    res = requests.post(url=url, json=body,headers=header)
    res_res = res.json()
    print(res_res)
    return res_res

def read_data(filename,sheetname):
    cases_list = []
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    for i in range(2,max_row+1,1):
        dict1 = dict(
        id = sheet.cell(row = i,column = 1).value,
        header = sheet.cell(row=i, column=5).value,
        url = sheet.cell(row=i, column=6).value,
        body = sheet.cell(row=i, column=7).value,
        expected = sheet.cell(row=i, column=8).value)
        cases_list.append(dict1)
    return cases_list

def write_data(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)

def execute_func(filename,sheetname):
    cases = read_data(filename,sheetname)
    # 通过read_data函数取出列表，通过for循环出里面的每个字典元素，然后通过key取出值
    for cs in cases:
        # Excel里面读取出来的数据，都是str类型！！！
        id = cs['id']            # 取出id
        url = cs['url']          # 取出接口地址
        header = cs['header']    # 取出请求头
        body = cs['body']        # 取出请求体
        expected = cs['expected']# 取出预期结果
        # eval(),内置函数,运行被字符串包裹着的python表达式，然后把引号去掉，取字符串里面的表达式
        header = eval(header)
        body = eval(body)
        expected = eval(expected)
        res = func(url=url,body=body,header=header)   # 调用func()函数发送请求
        expected_code = expected['code']  # 取出预期结果里面的code值
        real_code = res['code']           # 取出实际结果里面的code值
        print('预期结果的code为{}'.format(expected_code))
        print('实际结果的code为{}'.format(real_code))
        if expected_code == real_code:
            print('{}功能,第{}条用例通过！'.format(sheetname,id))
            print('*'*50)
            final_res = '通过'            # 执行结果
        else:
            print('{}功能,第{}条用例不通过！'.format(sheetname,id))
            print('*'*50)
            final_res = '不通过'          # 执行结果
        write_data(filename,sheetname,id+1,9,final_res)

execute_func('testcase_api_wuye.xlsx','login')
